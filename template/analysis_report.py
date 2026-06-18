#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""★営業日報2026年5月 分析レポート生成"""
import io, warnings
warnings.filterwarnings("ignore")
import numpy as np
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
import matplotlib.ticker as mticker
import matplotlib.colors as mcolors
from collections import defaultdict
import openpyxl
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

import matplotlib.font_manager as fm
_avail = {f.name for f in fm.fontManager.ttflist}
JP = next((f for f in ["Yu Gothic", "Meiryo", "MS Gothic"] if f in _avail), "sans-serif")
plt.rcParams.update({"font.family": JP, "axes.unicode_minus": False})

# ── ユーティリティ ──────────────────────────────
def to_buf(fig, dpi=130):
    buf = io.BytesIO()
    fig.savefig(buf, format="png", dpi=dpi, bbox_inches="tight",
                facecolor=fig.get_facecolor())
    buf.seek(0); plt.close(fig); return buf

# ── データ読み込み ──────────────────────────────
SRC = r"../data/★営業日報2026年5月.xlsx"
wb_src = openpyxl.load_workbook(SRC, data_only=True)
ws_d = wb_src["データ"]
ws_i = wb_src["商品別"]

rows = list(ws_d.iter_rows(min_row=1, max_row=54, values_only=True))

# 日次データ (col F=index5 〜 col AJ=index35)
COL0 = 5
date_raw  = rows[0][COL0:COL0+31]
hol_raw   = rows[1][COL0:COL0+31]
wday_raw  = rows[3][COL0:COL0+31]
sales_raw = rows[4][COL0:COL0+31]   # 純売上高

days, wdays, closed, d_sales = [], [], [], []
for d, h, w, s in zip(date_raw, hol_raw, wday_raw, sales_raw):
    if d is not None:
        days.append(d.day)
        wdays.append(w or "")
        closed.append(h == "休")
        d_sales.append(s or 0)

# 月次値（data_only=True 読み込み確認済み）
TOTAL     = 20_989_657
CASH      = 8_692_051
JCB       = 3_806_445
CHIBA     = 5_643_149
AQUA      = 567_876
PAYPAY    = 1_076_179
KAKEKIN   = 1_215_401
FOOD      = 14_975_700
DRINK     = 3_000_230
BAITEN    = 230_260
SONOTA    = 2_783_467
LUNCH_PAX = 2_520
DIN_PAX   = 1_534
LUNCH_AMT = 9_681_482
DIN_AMT   = 11_308_175

# 商品別集計
item_rows = list(ws_i.iter_rows(min_row=3, max_row=95, values_only=True))
food_s, drink_s = defaultdict(int), defaultdict(int)
for r in item_rows:
    if r[4] and r[7]:   food_s[r[4]]  += (r[7]  or 0)
    if r[8] and r[11]:  drink_s[r[8]] += (r[11] or 0)

top_food  = sorted(food_s.items(),  key=lambda x: x[1], reverse=True)[:8]
top_drink = sorted(drink_s.items(), key=lambda x: x[1], reverse=True)[:8]

# ── 共通スタイル関数 ─────────────────────────────
def clean_ax(ax):
    ax.spines[["top", "right"]].set_visible(False)
    ax.spines[["left", "bottom"]].set_edgecolor("#ddd")
    ax.tick_params(colors="#555", labelsize=8)

def y_man(ax):
    ax.yaxis.set_major_formatter(mticker.FuncFormatter(
        lambda v, _: f"{int(v/10000)}万"))

def x_man(ax):
    ax.xaxis.set_major_formatter(mticker.FuncFormatter(
        lambda v, _: f"{int(v/10000)}万"))

# ══════════════════════════════════════════════════
# ① 日次売上トレンド
# ══════════════════════════════════════════════════
def chart_1():
    fig, ax = plt.subplots(figsize=(13.5, 4.5), facecolor="white")

    bar_cols = ["#CFE2F3" if c else "#1E88E5" for c in closed]
    ax.bar(days, d_sales, color=bar_cols, width=0.72, zorder=2,
           edgecolor="white", lw=0.3)

    ox = [d for d, c in zip(days, closed) if not c]
    oy = [s for s, c in zip(d_sales, closed) if not c]
    ax.plot(ox, oy, "o-", color="#E53935", lw=1.8, markersize=4,
            zorder=3, label="売上推移")

    avg = np.mean(oy) if oy else 0
    ax.axhline(avg, color="#FF9800", lw=1.5, ls="--", alpha=0.85,
               label=f"稼働日平均: ¥{avg:,.0f}")

    for d, c in zip(days, closed):
        if c:
            ax.text(d, max(d_sales) * 0.02, "定休", ha="center",
                    va="bottom", fontsize=6, color="#aaa", rotation=90)

    ax.set_xticks(days)
    ax.set_xticklabels([f"{d}日\n({w})" for d, w in zip(days, wdays)],
                       fontsize=6.5)
    y_man(ax)
    ax.set_title("① 日次売上トレンド（2026年5月）", fontsize=13,
                 fontweight="bold", pad=10, color="#212121")
    ax.set_ylabel("純売上高（円）", fontsize=9, color="#555")
    clean_ax(ax)
    ax.grid(axis="y", color="#f0f0f0", lw=0.8, zorder=0)
    ax.set_ylim(0, max(d_sales) * 1.28)
    ax.legend(fontsize=9, loc="upper left", framealpha=0.8)
    fig.tight_layout()
    return to_buf(fig)

# ══════════════════════════════════════════════════
# ② 曜日別平均売上
# ══════════════════════════════════════════════════
def chart_2():
    order = ["月", "火", "水", "木", "金", "土", "日"]
    wmap = defaultdict(list)
    for s, w, c in zip(d_sales, wdays, closed):
        if not c and w:
            wmap[w].append(s)
    avgs   = [np.mean(wmap[w]) if wmap[w] else 0 for w in order]
    counts = [len(wmap[w]) for w in order]
    pal    = ["#90CAF9","#90CAF9","#90CAF9","#90CAF9",
              "#FFCC02","#4CAF50","#4CAF50"]

    fig, ax = plt.subplots(figsize=(7.5, 4.5), facecolor="white")
    bars = ax.bar(order, avgs, color=pal, width=0.6, zorder=2,
                  edgecolor="white", lw=0.3)
    max_avg = max(avgs) if avgs else 1
    for bar, avg, n in zip(bars, avgs, counts):
        if avg > 0:
            ax.text(bar.get_x() + bar.get_width() / 2,
                    avg + max_avg * 0.025,
                    f"{avg/10000:.1f}万\nn={n}",
                    ha="center", va="bottom",
                    fontsize=8.5, fontweight="bold", color="#333")
    y_man(ax)
    ax.set_title("② 曜日別 平均売上（稼働日のみ）", fontsize=12,
                 fontweight="bold", pad=10, color="#212121")
    ax.set_ylabel("平均純売上高（円）", fontsize=9, color="#555")
    clean_ax(ax)
    ax.grid(axis="y", color="#f0f0f0", lw=0.8)
    ax.set_ylim(0, max_avg * 1.32)
    fig.tight_layout()
    return to_buf(fig)

# ══════════════════════════════════════════════════
# ③ 支払い方法別構成
# ══════════════════════════════════════════════════
def chart_3():
    labels = ["現金", "JCB", "千葉銀行", "アクアコイン", "PayPay", "売掛金"]
    vals   = [CASH, JCB, CHIBA, AQUA, PAYPAY, KAKEKIN]
    colors = ["#1E88E5","#43A047","#FB8C00","#8E24AA","#E53935","#00ACC1"]

    fig, ax = plt.subplots(figsize=(6.5, 5.5), facecolor="white")
    ax.set_aspect("equal")
    _, _, autotexts = ax.pie(
        vals, colors=colors, startangle=90,
        autopct=lambda p: f"{p:.1f}%" if p > 2 else "",
        wedgeprops=dict(width=0.58, edgecolor="white", linewidth=2.5),
        pctdistance=0.78, counterclock=False
    )
    for at in autotexts:
        at.set_fontsize(8.5); at.set_fontweight("bold"); at.set_color("white")

    ax.add_patch(plt.Circle((0, 0), 0.42, color="white", zorder=10))
    ax.text(0, 0.12, "総売上高", ha="center", va="center",
            fontsize=8.5, color="#888")
    ax.text(0, -0.14, f"¥{TOTAL:,}", ha="center", va="center",
            fontsize=10.5, fontweight="bold", color="#212121")

    patches = [mpatches.Patch(color=c, label=f"{l}  ¥{v:,}")
               for c, l, v in zip(colors, labels, vals)]
    ax.legend(handles=patches, loc="lower center",
              bbox_to_anchor=(0.5, -0.2), ncol=2,
              fontsize=8, framealpha=0, labelspacing=0.5)
    ax.set_title("③ 支払い方法別 売上構成", fontsize=12,
                 fontweight="bold", pad=10, color="#212121")
    fig.tight_layout()
    return to_buf(fig)

# ══════════════════════════════════════════════════
# ④ カテゴリ別構成（ドーナツ＋横棒）
# ══════════════════════════════════════════════════
def chart_4():
    labels = ["FOOD", "DRINK", "売店", "その他"]
    vals   = [FOOD, DRINK, BAITEN, SONOTA]
    colors = ["#FF7043","#42A5F5","#66BB6A","#AB47BC"]
    total  = sum(vals)

    fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(11, 5), facecolor="white")

    # 左: ドーナツ
    ax1.set_aspect("equal")
    _, _, ats = ax1.pie(
        vals, colors=colors, startangle=90,
        autopct="%1.1f%%",
        wedgeprops=dict(width=0.55, edgecolor="white", linewidth=2.5),
        pctdistance=0.78, counterclock=False
    )
    for at in ats:
        at.set_fontsize(9); at.set_fontweight("bold"); at.set_color("white")
    ax1.add_patch(plt.Circle((0, 0), 0.45, color="white", zorder=10))
    ax1.text(0, 0.12, "総売上高", ha="center", fontsize=8.5, color="#888")
    ax1.text(0, -0.14, f"¥{total:,}", ha="center",
             fontsize=10, fontweight="bold", color="#212121")
    patches = [mpatches.Patch(color=c, label=f"{l}  ¥{v:,}")
               for c, l, v in zip(colors, labels, vals)]
    ax1.legend(handles=patches, loc="lower center",
               bbox_to_anchor=(0.5, -0.12), ncol=2,
               fontsize=8.5, framealpha=0)
    ax1.set_title("カテゴリ構成比", fontsize=10,
                  fontweight="bold", color="#212121", pad=8)

    # 右: 横棒
    y = np.arange(len(labels))
    ax2.barh(y, vals, color=colors, height=0.55,
             edgecolor="white", lw=0.5)
    ax2.set_yticks(y); ax2.set_yticklabels(labels, fontsize=10)
    for i, v in enumerate(vals):
        ax2.text(v + total * 0.012, i,
                 f"¥{v:,}  ({v/total*100:.1f}%)",
                 va="center", fontsize=8.5, color="#333")
    x_man(ax2)
    clean_ax(ax2)
    ax2.grid(axis="x", color="#f0f0f0", lw=0.8)
    ax2.set_xlim(0, max(vals) * 1.45)
    ax2.set_title("カテゴリ別 売上金額", fontsize=10,
                  fontweight="bold", color="#212121", pad=8)

    fig.suptitle("④ カテゴリ別 売上構成（FOOD / DRINK / 売店 / その他）",
                 fontsize=12, fontweight="bold", color="#212121", y=1.01)
    fig.tight_layout()
    return to_buf(fig)

# ══════════════════════════════════════════════════
# ⑤ 昼食 vs 夕食（複合グラフ）
# ══════════════════════════════════════════════════
def chart_5():
    labels = ["昼食", "夕食"]
    amts   = [LUNCH_AMT, DIN_AMT]
    pax    = [LUNCH_PAX, DIN_PAX]
    unit   = [a / p for a, p in zip(amts, pax)]
    colors = ["#FFB300", "#1E88E5"]

    fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(10, 4.8), facecolor="white")

    # 左: 売上金額
    x = np.arange(2)
    bars1 = ax1.bar(x, amts, color=colors, width=0.5,
                    edgecolor="white", lw=0.5)
    for bar, a in zip(bars1, amts):
        ax1.text(bar.get_x() + bar.get_width() / 2, a + max(amts) * 0.025,
                 f"¥{a:,}", ha="center",
                 fontsize=9, fontweight="bold", color="#333")
    ax1.set_xticks(x); ax1.set_xticklabels(labels, fontsize=11)
    y_man(ax1); clean_ax(ax1)
    ax1.grid(axis="y", color="#f0f0f0", lw=0.8)
    ax1.set_title("売上金額", fontsize=10, fontweight="bold", color="#212121")
    ax1.set_ylim(0, max(amts) * 1.3)

    # 右: 客数（棒）＋客単価（第2軸折れ線）
    ax2r = ax2.twinx()
    ax2.bar(x, pax, color=colors, width=0.5, alpha=0.8,
            edgecolor="white", lw=0.5)
    ax2r.plot(x, unit, "D-", color="#E53935", lw=2.2,
              markersize=9, zorder=5)
    for xi, (p, u) in enumerate(zip(pax, unit)):
        ax2.text(xi, p + max(pax) * 0.035, f"{p:,}名",
                 ha="center", fontsize=9, fontweight="bold", color="#333")
        ax2r.text(xi, u + max(unit) * 0.04, f"¥{u:,.0f}",
                  ha="center", fontsize=9,
                  fontweight="bold", color="#E53935")
    ax2.set_xticks(x); ax2.set_xticklabels(labels, fontsize=11)
    ax2.spines[["top"]].set_visible(False)
    ax2r.spines[["top", "left"]].set_visible(False)
    ax2.spines[["left", "bottom"]].set_edgecolor("#ddd")
    ax2.grid(axis="y", color="#f0f0f0", lw=0.8)
    ax2.set_ylabel("客数（名）", fontsize=9, color="#555")
    ax2r.set_ylabel("客単価（円）", fontsize=9, color="#E53935")
    ax2r.tick_params(axis="y", colors="#E53935", labelsize=8)
    ax2.tick_params(colors="#555", labelsize=8)
    ax2.set_title("客数 & 客単価", fontsize=10,
                  fontweight="bold", color="#212121")
    ax2.set_ylim(0, max(pax) * 1.38)
    ax2r.set_ylim(0, max(unit) * 1.5)

    fig.suptitle("⑤ 昼食 vs 夕食 — 売上・客数・客単価",
                 fontsize=12, fontweight="bold", color="#212121")
    fig.tight_layout()
    return to_buf(fig)

# ══════════════════════════════════════════════════
# ⑥ 売れ筋商品ランキング
# ══════════════════════════════════════════════════
def chart_6():
    fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(13.5, 5.5), facecolor="white")

    def _rank(ax, items, base_hex, title):
        names = [x[0] for x in items]
        vals  = [x[1] for x in items]
        y = np.arange(len(names))
        rgb = mcolors.to_rgb(base_hex)
        max_v = max(vals) if vals else 1
        bar_cols = [tuple(c * (0.38 + 0.62 * v / max_v)
                          + (1 - (0.38 + 0.62 * v / max_v)) * 0.93
                          for c in rgb) for v in vals]
        ax.barh(y, vals, color=bar_cols, height=0.65,
                edgecolor="white", lw=0.5)
        ax.set_yticks(y); ax.set_yticklabels(names, fontsize=8.5)
        ax.invert_yaxis()
        for i, v in enumerate(vals):
            ax.text(v + max_v * 0.015, i, f"¥{v:,}",
                    va="center", fontsize=8, color="#333")
        x_man(ax); clean_ax(ax)
        ax.grid(axis="x", color="#f0f0f0", lw=0.8)
        ax.set_xlim(0, max_v * 1.38)
        ax.set_title(title, fontsize=10.5, fontweight="bold",
                     color="#212121", pad=8)

    _rank(ax1, top_food,  "#FF7043", "FOOD 売上ランキング（月間累計 Top8）")
    _rank(ax2, top_drink, "#1E88E5", "DRINK 売上ランキング（月間累計 Top8）")

    fig.suptitle("⑥ 売れ筋商品ランキング",
                 fontsize=13, fontweight="bold", color="#212121")
    fig.tight_layout()
    return to_buf(fig)

# ── グラフ生成 ──────────────────────────────────
print("グラフ生成中...")
c1 = chart_1()
c2 = chart_2()
c3 = chart_3()
c4 = chart_4()
c5 = chart_5()
c6 = chart_6()
print("→ 6チャート完了")

# ══════════════════════════════════════════════════
# Excel レポート組み立て
# ══════════════════════════════════════════════════
print("Excelレポート生成中...")
wb = Workbook()

NAVY   = "1A237E"
BLUE   = "1565C0"
L_BLUE = "E3F2FD"
WHITE  = "FFFFFF"

def hdr_cell(ws, row, col, text, size=11, bold=True,
             fg="FFFFFF", bg="1A237E", align="center",
             end_col=None):
    if end_col:
        ws.merge_cells(start_row=row, start_column=col,
                       end_row=row, end_column=end_col)
    cell = ws.cell(row=row, column=col, value=text)
    cell.font = Font(name="Yu Gothic", size=size, bold=bold, color=fg)
    cell.alignment = Alignment(horizontal=align, vertical="center",
                               wrap_text=True)
    cell.fill = PatternFill("solid", fgColor=bg)
    return cell

def add_img(ws, buf, cell_anchor, w_px, h_px):
    img = XLImage(buf)
    img.width, img.height = w_px, h_px
    ws.add_image(img, cell_anchor)

# ────────────────────────────────────────
# シート0: 📋 サマリー
# ────────────────────────────────────────
ws0 = wb.active
ws0.title = "📋 サマリー"
ws0.sheet_view.showGridLines = False

col_widths = [2, 18, 18, 18, 18, 18, 18, 18, 18]
for i, w in enumerate(col_widths, 1):
    ws0.column_dimensions[get_column_letter(i)].width = w

# タイトル
hdr_cell(ws0, 1, 2, "【分析レポート】2026年5月 営業日報",
         16, bg=NAVY, end_col=9)
ws0.row_dimensions[1].height = 38

hdr_cell(ws0, 2, 2,
         "集計期間: 2026年5月1日〜5月31日　稼働日: 27日 / 定休: 4日",
         10, bg="283593", fg="BBCEFF", end_col=9)
ws0.row_dimensions[2].height = 22

# KPI セクション
hdr_cell(ws0, 4, 2, "■ 月間 KPI サマリー",
         11, bg=BLUE, end_col=9)
ws0.row_dimensions[4].height = 24

kpis = [
    ("総売上高",    TOTAL,          "¥#,##0"),
    ("純売上高",    19_097_801,     "¥#,##0"),
    ("FOOD売上",    FOOD,           "¥#,##0"),
    ("DRINK売上",   DRINK,          "¥#,##0"),
    ("現金売上",    CASH,           "¥#,##0"),
    ("クレジット",  JCB + CHIBA,    "¥#,##0"),
    ("昼食客数",    LUNCH_PAX,      '#,##0"名"'),
    ("夕食客数",    DIN_PAX,        '#,##0"名"'),
]
ws0.row_dimensions[5].height = 18
ws0.row_dimensions[6].height = 28
for i, (label, value, fmt) in enumerate(kpis):
    col = 2 + i
    lc = ws0.cell(row=5, column=col, value=label)
    lc.font = Font(name="Yu Gothic", size=8.5, color="555555")
    lc.alignment = Alignment(horizontal="center")
    lc.fill = PatternFill("solid", fgColor=L_BLUE)
    vc = ws0.cell(row=6, column=col, value=value)
    vc.font = Font(name="Yu Gothic", size=12, bold=True, color=BLUE)
    vc.alignment = Alignment(horizontal="center")
    vc.fill = PatternFill("solid", fgColor=L_BLUE)
    vc.number_format = fmt

# 分析シート一覧
hdr_cell(ws0, 8, 2, "■ 分析チャート 一覧",
         11, bg=BLUE, end_col=9)
ws0.row_dimensions[8].height = 24

nav = [
    ("① 日次売上トレンド",       "稼働日の日次推移・ピーク日・稼働日平均"),
    ("② 曜日別 平均売上",         "曜日ごとの平均売上（稼働日集計）"),
    ("③ 支払い方法別 売上構成",   "現金/JCB/千葉銀行/PayPay/アクアコイン/売掛金"),
    ("④ カテゴリ別 売上構成",     "FOOD/DRINK/売店/その他の比率と金額"),
    ("⑤ 昼食 vs 夕食 比較",      "売上・客数・客単価の昼夜比較"),
    ("⑥ 売れ筋商品ランキング",    "月間累計売上 FOOD/DRINK それぞれTop8"),
]
for i, (title, desc) in enumerate(nav):
    r = 10 + i
    bg = L_BLUE if i % 2 == 0 else WHITE
    tc = ws0.cell(row=r, column=2, value=title)
    tc.font = Font(name="Yu Gothic", size=10, bold=True, color=BLUE)
    tc.alignment = Alignment(horizontal="left", vertical="center")
    tc.fill = PatternFill("solid", fgColor=bg)
    ws0.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
    dc = ws0.cell(row=r, column=5, value=desc)
    dc.font = Font(name="Yu Gothic", size=9, color="555555")
    dc.alignment = Alignment(horizontal="left", vertical="center")
    dc.fill = PatternFill("solid", fgColor=bg)
    ws0.merge_cells(start_row=r, start_column=5, end_row=r, end_column=9)
    ws0.row_dimensions[r].height = 20

# ────────────────────────────────────────
# チャートシート生成
# ────────────────────────────────────────
chart_sheets = [
    ("①日次売上",       "① 日次売上トレンド（2026年5月）",                    c1, 990, 345),
    ("②曜日別",         "② 曜日別 平均売上",                                   c2, 570, 350),
    ("③支払方法",       "③ 支払い方法別 売上構成",                             c3, 495, 430),
    ("④カテゴリ",       "④ カテゴリ別 売上構成（FOOD / DRINK / 売店 / その他）", c4, 850, 390),
    ("⑤昼夜比較",       "⑤ 昼食 vs 夕食 — 売上・客数・客単価",               c5, 770, 375),
    ("⑥商品ランキング", "⑥ 売れ筋商品ランキング（月間累計 Top8）",             c6, 990, 435),
]

for sheet_name, title, chart_buf, img_w, img_h in chart_sheets:
    ws = wb.create_sheet(sheet_name)
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 130   # 広めに確保

    hdr_cell(ws, 1, 2, title, 13, bg=NAVY, end_col=16)
    ws.row_dimensions[1].height = 32
    ws.row_dimensions[2].height = 8   # 余白

    add_img(ws, chart_buf, "B3", img_w, img_h)

# ── 保存 ──────────────────────────────────────
OUT = r"../分析レポート_2026年5月.xlsx"
wb.save(OUT)
print(f"完了 → {OUT}")
