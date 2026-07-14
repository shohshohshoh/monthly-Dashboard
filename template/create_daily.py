#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
★営業日報yyyy年m月.xlsx → daily_yyyy_m.xlsx 変換スクリプト

使い方:
    python create_daily.py 2026 5
    → frontend/public/data/daily_2026_5.xlsx を出力

変換元（データシート横持ち形式）→ 変換先（縦持ち日次テーブル）
"""
import sys
from pathlib import Path
from datetime import datetime
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

ROOT    = Path(__file__).parent.parent
SRC_DIR = ROOT / "data"
OUT_DIR = ROOT / "frontend" / "public" / "data"

# データシートの行番号（固定レイアウト）
R_TEIKYU    = 2   # 定休（日曜=✗）
R_KYUJITSU  = 3   # 祝日FLG
R_YOUBI     = 4   # 曜日
R_JUN       = 5   # 純売上高
R_ZEIZEI    = 6   # 消費税
R_CASH      = 7   # 現金
R_TOTAL     = 10  # 総売上高
R_JCB       = 15  # JCB金額
R_CHIBA     = 16  # 千葉銀行金額
R_AQUA_S    = 20; R_AQUA_E = 24   # アクアコイン
R_PAY_S     = 25; R_PAY_E  = 34   # PayPay
R_FURU_S    = 35; R_FURU_E = 36   # ふるさと納税
R_URIKAKE   = 37  # 売掛金計
R_FOOD      = 41  # FOOD
R_DRINK     = 42  # DRINK
R_BAITEN    = 43  # 売店
R_OTHER     = 44  # その他
R_HIRU_KAK  = 47  # 昼食客数
R_YU_KAK    = 48  # 夕食客数
R_HIRU_AMT  = 50  # 昼食売上
R_YU_AMT    = 51  # 夕食売上


def _int(val):
    try:
        return int(val) if val is not None else 0
    except (ValueError, TypeError):
        return 0


def _read_grid(ws, max_row, max_col):
    """ws を1回だけ前方向に読み、grid[row][col-1] (1-indexed row) で参照できるリストに変換する。
    read_only ワークシートは逆方向シークができず、cell()でのランダムアクセスは
    先頭から再走査するため極端に遅くなる。1回のforward iterationでキャッシュしておく。"""
    grid = [None]
    for row in ws.iter_rows(min_row=1, max_row=max_row, max_col=max_col, values_only=True):
        grid.append(row)
    return grid


def _gv(grid, row, col):
    if row >= len(grid) or grid[row] is None or col > len(grid[row]):
        return None
    return grid[row][col - 1]


def _sum(grid, r_start, r_end, col):
    return sum(_int(_gv(grid, r, col)) for r in range(r_start, r_end + 1))


# 元Excelの書式崩れ（余分な行・列への書式設定等）で ws.max_row / ws.max_column が
# 異常に大きくなるケースへの耐性のため、実際に必要な範囲を明示的な上限で区切る。
_DATE_COL_START = 6
_DATE_COL_MAX   = _DATE_COL_START + 40   # 1か月は最大31日、余裕を見て+40列まで走査
_SHOHIN_MAX_ROW = 2000                   # 商品別データは通常でも数百行程度


def create_daily(year: int, month: int) -> Path:
    src = SRC_DIR / f"★営業日報{year}年{month}月.xlsx"
    if not src.exists():
        raise FileNotFoundError(f"{src} が見つかりません")

    out = OUT_DIR / f"daily_{year}_{month}.xlsx"
    OUT_DIR.mkdir(parents=True, exist_ok=True)

    # read_only=True: 「データ」「商品別」以外のシートに書式崩れ等でメモリを
    # 大量消費する内容が残っていても、アクセスしない限り読み込まれない。
    wb_src    = openpyxl.load_workbook(str(src), data_only=True, read_only=True)
    ws_data   = wb_src["データ"]
    ws_shohin = wb_src["商品別"]

    grid_data = _read_grid(ws_data, R_YU_AMT, _DATE_COL_MAX)

    # 行1 でデータ列を特定（列F以降、datetime値）
    date_cols = []
    for col in range(_DATE_COL_START, _DATE_COL_MAX + 1):
        val = _gv(grid_data, 1, col)
        if val is None:
            break
        if isinstance(val, datetime) and val.year == year and val.month == month:
            date_cols.append((col, val))

    # ── 日次データ ──
    KEYS = ["日付", "曜日", "定休", "祝日",
            "純売上高", "消費税", "総売上高",
            "現金", "JCB", "千葉銀行",
            "アクアコイン", "PayPay", "ふるさと納税", "売掛金",
            "FOOD", "DRINK", "売店", "その他",
            "昼食客数", "夕食客数", "昼食売上", "夕食売上"]

    daily_rows = []
    for col, dt in date_cols:
        teikyu   = _gv(grid_data, R_TEIKYU, col) or None
        kyujitsu = _gv(grid_data, R_KYUJITSU, col) or None

        total = _int(_gv(grid_data, R_TOTAL, col))
        if total == 0:
            teikyu = "休"

        row = {
            "日付":         dt,
            "曜日":         _gv(grid_data, R_YOUBI, col) or "",
            "定休":         teikyu,
            "祝日":         kyujitsu,
            "純売上高":     _int(_gv(grid_data, R_JUN,      col)),
            "消費税":       _int(_gv(grid_data, R_ZEIZEI,   col)),
            "総売上高":     _int(_gv(grid_data, R_TOTAL,    col)),
            "現金":         _int(_gv(grid_data, R_CASH,     col)),
            "JCB":          _int(_gv(grid_data, R_JCB,      col)),
            "千葉銀行":     _int(_gv(grid_data, R_CHIBA,    col)),
            "アクアコイン": _sum(grid_data, R_AQUA_S, R_AQUA_E, col),
            "PayPay":       _sum(grid_data, R_PAY_S,  R_PAY_E,  col),
            "ふるさと納税": _sum(grid_data, R_FURU_S, R_FURU_E, col),
            "売掛金":       _int(_gv(grid_data, R_URIKAKE,  col)),
            "FOOD":         _int(_gv(grid_data, R_FOOD,     col)),
            "DRINK":        _int(_gv(grid_data, R_DRINK,    col)),
            "売店":         _int(_gv(grid_data, R_BAITEN,   col)),
            "その他":       _int(_gv(grid_data, R_OTHER,    col)),
            "昼食客数":     _int(_gv(grid_data, R_HIRU_KAK, col)),
            "夕食客数":     _int(_gv(grid_data, R_YU_KAK,   col)),
            "昼食売上":     _int(_gv(grid_data, R_HIRU_AMT, col)),
            "夕食売上":     _int(_gv(grid_data, R_YU_AMT,   col)),
        }
        daily_rows.append(row)

    # ── 商品別データ ──
    SHOHIN_KEYS = ["日付", "日", "曜日", "順位",
                   "F商品名", "F単価", "F数量", "F金額",
                   "D商品名", "D単価", "D数量", "D金額"]

    grid_shohin = _read_grid(ws_shohin, _SHOHIN_MAX_ROW, 12)

    shohin_rows = []
    for r in range(3, _SHOHIN_MAX_ROW + 1):
        vals = [_gv(grid_shohin, r, c) for c in range(1, 13)]
        if vals[0] is None:
            continue
        dt = vals[0]
        if not isinstance(dt, datetime):
            continue
        shohin_rows.append({
            "日付":    dt.strftime("%Y/%m/%d"),
            "日":      dt.day,
            "曜日":    vals[2] or "",
            "順位":    vals[3] or "",
            "F商品名": vals[4],  "F単価": int(vals[5]  or 0),
            "F数量":   int(vals[6]  or 0), "F金額": int(vals[7]  or 0),
            "D商品名": vals[8],  "D単価": int(vals[9]  or 0),
            "D数量":   int(vals[10] or 0), "D金額": int(vals[11] or 0),
        })

    # ── Excel出力 ──
    _TABLE_STYLE = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False, showLastColumn=False,
        showRowStripes=True, showColumnStripes=False,
    )

    wb_out = openpyxl.Workbook()

    # 日次データシート（テーブル形式、ヘッダーは行1から）
    ws1 = wb_out.active
    ws1.title = f"{year}年{month}月_日次データ"
    for ci, h in enumerate(KEYS, 1):
        ws1.cell(1, ci, h)
    for ri, row in enumerate(daily_rows, 2):
        for ci, key in enumerate(KEYS, 1):
            cell = ws1.cell(ri, ci, row[key])
            if key == "日付":
                cell.number_format = "yyyy/mm/dd"

    last_col1 = get_column_letter(len(KEYS))
    last_row1  = max(1 + len(daily_rows), 2)  # テーブルは最低2行（ヘッダー+1）
    tbl1 = Table(displayName="DailyData", ref=f"A1:{last_col1}{last_row1}")
    tbl1.tableStyleInfo = _TABLE_STYLE
    ws1.add_table(tbl1)

    # 商品別シート（テーブル形式、ヘッダーは行1から）
    ws2 = wb_out.create_sheet("商品別")
    for ci, h in enumerate(SHOHIN_KEYS, 1):
        ws2.cell(1, ci, h)
    for ri, row in enumerate(shohin_rows, 2):
        for ci, key in enumerate(SHOHIN_KEYS, 1):
            ws2.cell(ri, ci, row[key])

    last_col2 = get_column_letter(len(SHOHIN_KEYS))
    last_row2  = max(1 + len(shohin_rows), 2)
    tbl2 = Table(displayName="Shohin", ref=f"A1:{last_col2}{last_row2}")
    tbl2.tableStyleInfo = _TABLE_STYLE
    ws2.add_table(tbl2)

    wb_out.save(str(out))
    print(f"保存: {out}  ({len(daily_rows)}日, 商品別{len(shohin_rows)}行)")
    return out


if __name__ == "__main__":
    if len(sys.argv) != 3:
        print("使い方: python create_daily.py <year> <month>")
        sys.exit(1)
    create_daily(int(sys.argv[1]), int(sys.argv[2]))
