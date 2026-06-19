#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
営業日報 Excel → サマリーExcel 変換スクリプト

使い方:
    python create_summary.py 2026 5
    → frontend/public/data/summary_2026_5.xlsx を出力

データシート セル対応表（E列 = 2026年5月累計）:
  行5  : 純売上高      行10 : 総売上高
  行7  : 現金          行15 : JCB
  行16 : 千葉銀行      行20 : アクアコイン
  行25 : PayPay        行37 : 売掛金
  行41 : FOOD          行42 : DRINK
  行43 : 売店          行44 : その他
  行47 : 昼食客数      行48 : 夕食客数
  行50 : 昼食売上金額  行51 : 夕食売上金額
  日次データ: 行5 F列〜（COL0=6, 最大31列）
"""
import sys
import argparse
from pathlib import Path
from collections import defaultdict
import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter

ROOT     = Path(__file__).parent.parent
DATA_DIR = ROOT / "data"
OUT_DIR  = ROOT / "frontend" / "public" / "data"

# ── スタイル定数 ──────────────────────────────────────────────
HDR_FILL  = PatternFill("solid", fgColor="08123A")
HDR_FONT  = Font(bold=True, color="00B4FF", size=10)
VAL_FONT  = Font(color="D0E8FF", size=10)
SUB_FONT  = Font(color="4A6A8A", size=9)
ROW_FILL  = PatternFill("solid", fgColor="030718")
ALT_FILL  = PatternFill("solid", fgColor="04091F")
THIN      = Side(style="thin", color="0F2D4A")
BORDER    = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER    = Alignment(horizontal="center", vertical="center")
RIGHT     = Alignment(horizontal="right",  vertical="center")
LEFT      = Alignment(horizontal="left",   vertical="center")


def _h(ws, row, col, text):
    """ヘッダーセル"""
    c = ws.cell(row=row, column=col, value=text)
    c.font = HDR_FONT; c.fill = HDR_FILL
    c.alignment = CENTER; c.border = BORDER


def _v(ws, row, col, value, fmt=None, align=RIGHT):
    """値セル"""
    c = ws.cell(row=row, column=col, value=value)
    c.font = VAL_FONT; c.fill = ROW_FILL if row % 2 == 0 else ALT_FILL
    c.alignment = align; c.border = BORDER
    if fmt: c.number_format = fmt


def _s(ws, row, col, text):
    """サブラベルセル"""
    c = ws.cell(row=row, column=col, value=text)
    c.font = SUB_FONT; c.fill = HDR_FILL
    c.alignment = LEFT; c.border = BORDER


# ── Excelデータ読み込み ──────────────────────────────────────
def load_data(year: int, month: int) -> dict:
    filename = f"★営業日報{year}年{month}月.xlsx"
    path     = DATA_DIR / filename
    if not path.exists():
        raise FileNotFoundError(f"ファイルが見つかりません: {path}")

    wb  = openpyxl.load_workbook(path, data_only=True)
    ws  = wb["データ"]

    def e(row): return int(ws.cell(row=row, column=5).value or 0)

    # ── 集計値（E列固定）
    total_sales = e(10)   # 総売上高（税込）
    net_sales   = e(5)    # 純売上高
    cash        = e(7)    # 現金
    jcb         = e(15)   # JCB
    chiba       = e(16)   # 千葉銀行
    aqua        = e(20)   # アクアコイン
    paypay      = e(25)   # PayPay
    kake        = e(37)   # 売掛金
    food        = e(41)   # FOOD
    drink       = e(42)   # DRINK
    baiten      = e(43)   # 売店
    sonota      = e(44)   # その他
    lunch_pax   = e(47)   # 昼食客数
    dinner_pax  = e(48)   # 夕食客数
    lunch_amt   = e(50)   # 昼食売上金額
    dinner_amt  = e(51)   # 夕食売上金額

    # ── 日次データ（行1〜5, F列〜）
    COL0   = 6   # F列
    rows_d = list(ws.iter_rows(min_row=1, max_row=5,
                               min_col=COL0, max_col=COL0+30,
                               values_only=True))
    date_r  = rows_d[0]  # 行1: 日付
    hol_r   = rows_d[1]  # 行2: 休日FLG
    wday_r  = rows_d[3]  # 行4: 曜日
    sales_r = rows_d[4]  # 行5: 純売上高

    daily = []
    for d, h, w, s in zip(date_r, hol_r, wday_r, sales_r):
        if d is not None:
            daily.append({
                "day":       d.day,
                "weekday":   w or "",
                "is_closed": h == "休",
                "sales":     int(s or 0),
            })

    working_days = sum(1 for r in daily if not r["is_closed"])
    closed_days  = sum(1 for r in daily if r["is_closed"])

    # ── 商品別Top6
    ws_item    = wb["商品別"]
    food_map   = defaultdict(int)
    drink_map  = defaultdict(int)
    for r in ws_item.iter_rows(min_row=3, max_row=95, values_only=True):
        if r[4] and r[7]:  food_map[r[4]]  += int(r[7]  or 0)
        if r[8] and r[11]: drink_map[r[8]] += int(r[11] or 0)

    top_food  = sorted(food_map.items(),  key=lambda x: x[1], reverse=True)[:6]
    top_drink = sorted(drink_map.items(), key=lambda x: x[1], reverse=True)[:6]

    return dict(
        year=year, month=month, filename=filename,
        working_days=working_days, closed_days=closed_days,
        total_sales=total_sales, net_sales=net_sales,
        cash=cash, jcb=jcb, chiba=chiba, aqua=aqua, paypay=paypay, kake=kake,
        food=food, drink=drink, baiten=baiten, sonota=sonota,
        lunch_pax=lunch_pax, dinner_pax=dinner_pax,
        lunch_amt=lunch_amt, dinner_amt=dinner_amt,
        daily=daily, top_food=top_food, top_drink=top_drink,
    )


# ── サマリーExcel書き出し ────────────────────────────────────
def write_summary_excel(d: dict, out_path: Path):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)   # デフォルトシート削除

    # ── シート1: KPIサマリー ──────────────────────────────
    ws1 = wb.create_sheet("KPIサマリー")
    ws1.sheet_view.showGridLines = False
    ws1.column_dimensions["A"].width = 18
    ws1.column_dimensions["B"].width = 22
    ws1.column_dimensions["C"].width = 16

    # タイトル行
    ws1.merge_cells("A1:C1")
    t = ws1["A1"]
    t.value = f"{d['year']}年{d['month']}月 営業日報 サマリー"
    t.font      = Font(bold=True, color="00B4FF", size=13)
    t.fill      = HDR_FILL
    t.alignment = CENTER
    t.border    = BORDER
    ws1.row_dimensions[1].height = 28

    # ── 基本情報
    _h(ws1, 3, 1, "項目"); _h(ws1, 3, 2, "値"); _h(ws1, 3, 3, "備考")
    info = [
        ("対象ファイル",   d["filename"],         ""),
        ("集計年月",        f"{d['year']}年{d['month']}月", ""),
        ("稼働日数",        d["working_days"],     "日"),
        ("定休日数",        d["closed_days"],      "日"),
    ]
    for i, (lbl, val, note) in enumerate(info, start=4):
        _s(ws1, i, 1, lbl)
        _v(ws1, i, 2, val,  align=CENTER)
        _v(ws1, i, 3, note, align=CENTER)

    # ── KPI
    ws1.merge_cells("A9:C9")
    s = ws1["A9"]
    s.value = "■ KPI"
    s.font = Font(bold=True, color="00F5A0", size=10)
    s.fill = HDR_FILL; s.alignment = LEFT; s.border = BORDER

    _h(ws1, 10, 1, "指標"); _h(ws1, 10, 2, "金額（円）"); _h(ws1, 10, 3, "備考")
    kpi = [
        ("総売上高（税込）", d["total_sales"],   ""),
        ("純売上高",          d["net_sales"],     "消費税除く"),
        ("FOOD売上",          d["food"],          ""),
        ("DRINK売上",         d["drink"],         ""),
        ("売店売上",          d["baiten"],        ""),
        ("その他売上",        d["sonota"],        ""),
        ("昼食売上金額",      d["lunch_amt"],     ""),
        ("夕食売上金額",      d["dinner_amt"],    ""),
        ("昼食客数",          d["lunch_pax"],     "名"),
        ("夕食客数",          d["dinner_pax"],    "名"),
    ]
    for i, (lbl, val, note) in enumerate(kpi, start=11):
        _s(ws1, i, 1, lbl)
        _v(ws1, i, 2, val, fmt='#,##0' if isinstance(val, int) else None)
        _v(ws1, i, 3, note, align=CENTER)

    # ── 支払方法別
    ws1.merge_cells("A22:C22")
    s = ws1["A22"]
    s.value = "■ 支払方法別"
    s.font = Font(bold=True, color="00F5A0", size=10)
    s.fill = HDR_FILL; s.alignment = LEFT; s.border = BORDER

    _h(ws1, 23, 1, "支払方法"); _h(ws1, 23, 2, "金額（円）"); _h(ws1, 23, 3, "構成比")
    payment = [
        ("現金",         d["cash"]),
        ("JCB",          d["jcb"]),
        ("千葉銀行",     d["chiba"]),
        ("アクアコイン", d["aqua"]),
        ("PayPay",       d["paypay"]),
        ("売掛金",       d["kake"]),
    ]
    total_p = sum(v for _, v in payment)
    for i, (lbl, val) in enumerate(payment, start=24):
        _s(ws1, i, 1, lbl)
        _v(ws1, i, 2, val, fmt='#,##0')
        pct = val / total_p * 100 if total_p else 0
        _v(ws1, i, 3, round(pct, 1), fmt='0.0"%"')

    # ── シート2: 日次売上 ──────────────────────────────────
    ws2 = wb.create_sheet("日次売上")
    ws2.sheet_view.showGridLines = False
    for col, w in zip("ABCDE", [8, 8, 12, 10, 12]):
        ws2.column_dimensions[get_column_letter(
            "ABCDE".index(col)+1)].width = w

    ws2.merge_cells("A1:E1")
    t = ws2["A1"]
    t.value = f"{d['year']}年{d['month']}月 日次売上"
    t.font = Font(bold=True, color="00B4FF", size=12)
    t.fill = HDR_FILL; t.alignment = CENTER; t.border = BORDER

    for col, lbl in enumerate(["日", "曜日", "定休", "純売上高（円）", "備考"], 1):
        _h(ws2, 2, col, lbl)

    for i, row in enumerate(d["daily"], start=3):
        _v(ws2, i, 1, row["day"],      fmt='0',      align=CENTER)
        _v(ws2, i, 2, row["weekday"],               align=CENTER)
        _v(ws2, i, 3, "休" if row["is_closed"] else "", align=CENTER)
        _v(ws2, i, 4, row["sales"],    fmt='#,##0')
        _v(ws2, i, 5, "", align=CENTER)

    # 合計行
    last = 3 + len(d["daily"])
    total_daily = sum(r["sales"] for r in d["daily"])
    ws2.merge_cells(f"A{last}:C{last}")
    _h(ws2, last, 1, "合 計")
    _v(ws2, last, 4, total_daily, fmt='#,##0')
    _v(ws2, last, 5, "", align=CENTER)

    # ── シート3: 商品別Top ─────────────────────────────────
    ws3 = wb.create_sheet("商品別Top")
    ws3.sheet_view.showGridLines = False
    ws3.column_dimensions["A"].width = 6
    ws3.column_dimensions["B"].width = 20
    ws3.column_dimensions["C"].width = 14
    ws3.column_dimensions["D"].width = 6
    ws3.column_dimensions["E"].width = 20
    ws3.column_dimensions["F"].width = 14

    ws3.merge_cells("A1:C1")
    t = ws3["A1"]
    t.value = "FOOD Top6"
    t.font = Font(bold=True, color="FF9F1C", size=11)
    t.fill = HDR_FILL; t.alignment = CENTER; t.border = BORDER

    ws3.merge_cells("D1:F1")
    t = ws3["D1"]
    t.value = "DRINK Top6"
    t.font = Font(bold=True, color="00B4FF", size=11)
    t.fill = HDR_FILL; t.alignment = CENTER; t.border = BORDER

    for col, lbl in enumerate(["順位", "商品名", "売上金額（円）"], 1):
        _h(ws3, 2, col, lbl)
    for col, lbl in enumerate(["順位", "商品名", "売上金額（円）"], 4):
        _h(ws3, 2, col, lbl)

    for i, (name, amt) in enumerate(d["top_food"], start=3):
        _v(ws3, i, 1, i-2,    align=CENTER)
        _v(ws3, i, 2, name,   align=LEFT)
        _v(ws3, i, 3, amt,    fmt='#,##0')
    for i, (name, amt) in enumerate(d["top_drink"], start=3):
        _v(ws3, i, 4, i-2,    align=CENTER)
        _v(ws3, i, 5, name,   align=LEFT)
        _v(ws3, i, 6, amt,    fmt='#,##0')

    # ── 保存
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)


def main():
    parser = argparse.ArgumentParser(description="営業日報 → サマリーExcel")
    parser.add_argument("year",  type=int, help="年（例: 2026）")
    parser.add_argument("month", type=int, help="月（例: 5）")
    args = parser.parse_args()

    print(f"読み込み中: ★営業日報{args.year}年{args.month}月.xlsx")
    d = load_data(args.year, args.month)

    out_path = OUT_DIR / f"summary_{args.year}_{args.month}.xlsx"
    write_summary_excel(d, out_path)
    print(f"出力完了: {out_path}")


if __name__ == "__main__":
    main()
