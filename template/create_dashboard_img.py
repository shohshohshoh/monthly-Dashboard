#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
日次データ → 1枚ダッシュボード PNG 生成スクリプト

使い方:
    python create_dashboard_img.py 2026 5
    → frontend/public/data/dashboard_2026_5.png を出力

レイアウト:
  [KPI×5                                                        ]
  [① 日次売上トレンド（8/12）          ] [④ カテゴリ別構成（4/12）]
  [② 曜日別（4/12）][⑦F（4/12）][⑦D（4/12）                   ]
  [⑧ 客単価日次推移（8/12）            ] [③ 支払方法別（4/12）   ]
"""
import argparse
from pathlib import Path
from collections import defaultdict
import numpy as np
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.gridspec as gridspec
import matplotlib.font_manager as fm
import openpyxl

ROOT    = Path(__file__).parent.parent
IN_DIR  = ROOT / "frontend" / "public" / "data"
OUT_DIR = ROOT / "frontend" / "public" / "data"

def _setup_jp_font():
    import glob as _g

    # 0. ビルド時にダウンロードした BIZ UDGothic Bold を優先（Render 環境）
    _dl_dir  = Path(__file__).parent / "fonts"
    _biz_bold = _dl_dir / "BIZUDGothic-Bold.ttf"
    _biz_reg  = _dl_dir / "BIZUDGothic-Regular.ttf"
    if _biz_bold.exists():
        fm.fontManager.addfont(str(_biz_bold))
        if _biz_reg.exists():
            fm.fontManager.addfont(str(_biz_reg))
        return fm.FontProperties(fname=str(_biz_bold)).get_name()

    # 1. japanize_matplotlib 同梱フォントを直接登録
    try:
        import japanize_matplotlib as _jm
        _pkg = Path(_jm.__file__).parent
        _fonts = sorted(_pkg.glob("**/*.ttf")) + sorted(_pkg.glob("**/*.otf"))
        if _fonts:
            _fp = str(_fonts[0])
            fm.fontManager.addfont(_fp)
            return fm.FontProperties(fname=_fp).get_name()
    except Exception:
        pass

    # 2. システムフォントを全件スキャンして登録（apt-get install fonts-ipafont 後）
    _sys = (
        _g.glob("/usr/share/fonts/**/*.ttf", recursive=True) +
        _g.glob("/usr/share/fonts/**/*.otf", recursive=True)
    )
    for _f in _sys:
        try:
            fm.fontManager.addfont(_f)
        except Exception:
            pass
    _ipa = next((f for f in _sys if "ipa" in f.lower()), None)
    if _ipa:
        try:
            return fm.FontProperties(fname=_ipa).get_name()
        except Exception:
            pass

    # 3. Windows ローカル環境フォント
    _avail = {f.name for f in fm.fontManager.ttflist}
    return next(
        (f for f in ["Yu Gothic", "Meiryo", "MS Gothic"]
         if f in _avail),
        "sans-serif"
    )

_JP = _setup_jp_font()
plt.rcParams.update({
    "font.family":        _JP,
    "axes.unicode_minus": False,
    "font.weight":        "bold",
    "axes.titleweight":   "bold",
    "axes.labelweight":   "bold",
})

C = {
    "bg":   "#f0f4f8", "card": "#ffffff",
    "c1":   "#0284c7", "c2":   "#7c3aed",
    "c3":   "#059669", "c4":   "#d97706",
    "c5":   "#e11d48", "c6":   "#ca8a04",
    "c7":   "#0891b2", "c8":   "#be123c",
    "grid": "#cbd5e1", "text": "#0f172a", "sub":  "#475569",
}


# ══════════════════════════════════════════════════════════════
# データ読み込み
# ══════════════════════════════════════════════════════════════
_WEEKDAY_JP = ["月", "火", "水", "木", "金", "土", "日"]


def load_data(year, month):
    path = IN_DIR / f"daily_{year}_{month}.xlsx"
    if not path.exists():
        raise FileNotFoundError(f"データファイルが見つかりません: {path}")
    wb  = openpyxl.load_workbook(path)
    ws1 = wb.active
    # フォーマット: ヘッダー行1、データ行2〜
    keys = [
        "日付", "曜日", "定休", "祝日",
        "純売上高", "消費税", "総売上高",
        "現金", "JCB", "千葉銀行", "アクアコイン", "PayPay", "ふるさと納税", "売掛金",
        "FOOD", "DRINK", "売店", "その他",
        "昼食客数", "夕食客数", "昼食売上", "夕食売上",
    ]
    daily = []
    for r in range(2, ws1.max_row + 1):
        v = [ws1.cell(row=r, column=c).value for c in range(1, len(keys) + 1)]
        if v[0] is None or str(v[0]).startswith("合"):
            continue
        row = dict(zip(keys, v))
        for k in keys[4:]:  # 純売上高 以降を数値化
            row[k] = int(row[k] or 0)
        # 日付から日を導出
        dt = row["日付"]
        if hasattr(dt, "day"):
            row["日"] = dt.day
        else:
            from datetime import datetime as _dt
            try:
                d = _dt.strptime(str(row["日付"]), "%Y/%m/%d")
                row["日"] = d.day
            except Exception:
                row["日"] = 0
        daily.append(row)

    ws2    = wb["商品別"]
    s_keys = ["日付", "日", "曜日", "順位",
              "F商品名", "F単価", "F数量", "F金額",
              "D商品名", "D単価", "D数量", "D金額"]
    shohin = []
    for r in range(2, ws2.max_row + 1):  # 新フォーマット: データ行2〜
        v = [ws2.cell(row=r, column=c).value for c in range(1, 13)]
        if v[0] is None or str(v[0]).startswith("合"):
            continue
        row = dict(zip(s_keys, v))
        for k in ["F単価", "F数量", "F金額", "D単価", "D数量", "D金額"]:
            row[k] = int(row[k] or 0)
        shohin.append(row)
    return daily, shohin


# ══════════════════════════════════════════════════════════════
# 描画ヘルパー
# ══════════════════════════════════════════════════════════════
def dark_ax(ax):
    ax.set_facecolor(C["card"])
    ax.tick_params(colors=C["text"], labelsize=8.5, length=2, pad=3)
    ax.spines["top"].set_visible(False)
    ax.spines["right"].set_visible(False)
    ax.spines["left"].set_edgecolor(C["grid"])
    ax.spines["bottom"].set_edgecolor(C["grid"])
    ax.grid(color=C["grid"], ls="--", lw=0.5, alpha=0.5, axis="y", zorder=0)


def neon_line(ax, xs, ys, color, lw=1.8, ms=3.5, marker="o", label="", zorder=3):
    ax.plot(xs, ys, color=color, lw=lw, marker=marker, ms=ms,
            markerfacecolor=C["bg"], markeredgecolor=color,
            markeredgewidth=1.4, label=label, solid_capstyle="round", zorder=zorder)


def sub_title(ax, text, color=None):
    ax.text(0.5, 0.975, text, transform=ax.transAxes,
            ha="center", va="top",
            color=color or C["text"], fontsize=12.5,
            fontweight="bold",
            bbox=dict(boxstyle="round,pad=0.25",
                      facecolor=C["bg"], edgecolor="none", alpha=0.72),
            zorder=10)


# ══════════════════════════════════════════════════════════════
# KPI カード
# ══════════════════════════════════════════════════════════════
def draw_kpi(ax, label, value, color):
    ax.set_facecolor(C["bg"])
    ax.set_xlim(0, 1); ax.set_ylim(0, 1)
    ax.axis("off")
    rect = plt.Rectangle((0.02, 0.06), 0.96, 0.90,
                          edgecolor=color, facecolor=C["card"],
                          linewidth=2.5, zorder=2, clip_on=False)
    ax.add_patch(rect)
    ax.text(0.10, 0.50, label, ha="left",  va="center",
            fontsize=18, color=color, fontweight="bold", zorder=3)
    ax.text(0.90, 0.50, value, ha="right", va="center",
            fontsize=18, color=color, fontweight="bold", zorder=3)


# ══════════════════════════════════════════════════════════════
# X軸ラベル（日＋曜日）共通ヘルパー
# ══════════════════════════════════════════════════════════════
def set_day_xticks(ax, op_days):
    """op_days: 稼働日のdictリスト（"日"と"曜日"キーを含む）"""
    xs     = [d["日"] for d in op_days]
    labels = [f"{d['日']}\n{d['曜日']}" for d in op_days]
    ax.set_xticks(xs)
    ax.set_xticklabels(labels, fontsize=7.8, fontweight="bold")
    ax.tick_params(axis="x", length=3, pad=1)


# ══════════════════════════════════════════════════════════════
# ① 日次売上トレンド
# ══════════════════════════════════════════════════════════════
def plot_c1(ax, daily):
    op     = [d for d in daily if d["定休"] != "休"]
    xs     = [d["日"] for d in op]
    total  = [d["総売上高"] / 10000 for d in op]
    lunch  = [d["昼食売上"]  / 10000 for d in op]
    dinner = [d["夕食売上"]  / 10000 for d in op]
    avg    = np.mean(total)
    max_i  = int(np.argmax(total))

    dark_ax(ax)
    ax.fill_between(xs, total,  alpha=0.28, color=C["c1"])
    ax.fill_between(xs, lunch,  alpha=0.22, color=C["c4"])
    ax.fill_between(xs, dinner, alpha=0.22, color=C["c2"])

    neon_line(ax, xs, total,  C["c1"], lw=2.0, ms=3,   label="総売上高")
    neon_line(ax, xs, lunch,  C["c4"], lw=1.4, ms=2.5, marker="^", label="昼食売上")
    neon_line(ax, xs, dinner, C["c2"], lw=1.4, ms=2.5, marker="s", label="夕食売上")
    for w, a in [(3, 0.06), (1.5, 0.12)]:
        ax.axhline(avg, color=C["c3"], lw=w, alpha=a, ls="--")
    ax.axhline(avg, color=C["c3"], lw=0.9, ls="--", alpha=0.65,
               label=f"平均 {avg:.0f}万")

    ax.text(xs[max_i], total[max_i] + max(total)*0.04,
            f"MAX\n{total[max_i]:.0f}万",
            ha="center", fontsize=8, color=C["c4"], fontweight="bold",
            bbox=dict(boxstyle="round,pad=0.2", fc=C["bg"], ec=C["c4"],
                      alpha=0.8, lw=0.8))

    ax.set_xlim(min(xs) - 0.5, max(xs) + 0.5)
    ax.set_ylim(0, max(total) + 15)
    ax.set_ylabel("売上（万円）", color=C["sub"], fontsize=9, fontweight="bold")
    ax.legend(facecolor=C["card"], edgecolor=C["grid"],
              labelcolor=C["text"], fontsize=8.5, framealpha=0.85,
              loc="upper left", borderpad=0.5)
    set_day_xticks(ax, op)
    sub_title(ax, "① 日次売上トレンド")


# ══════════════════════════════════════════════════════════════
# ④ カテゴリ別構成（円グラフ・上右エリア）
# ══════════════════════════════════════════════════════════════
def plot_c4(ax, daily):
    cats   = ["FOOD", "DRINK", "その他"]
    colors = [C["c4"], C["c1"], C["c2"]]
    raw    = {k: sum(d[k] for d in daily) for k in ["FOOD", "DRINK", "売店", "その他"]}
    vals   = [raw["FOOD"], raw["DRINK"], raw["売店"] + raw["その他"]]
    total  = sum(vals)

    ax.set_facecolor(C["bg"])
    ax.set_aspect("equal")

    R = 1.18
    ax.pie(vals, colors=colors, startangle=90, radius=R,
           wedgeprops=dict(edgecolor=C["bg"], linewidth=2.5),
           counterclock=False)

    # 全スライス内部にラベル
    start_deg = 90.0
    for lbl, v, co in zip(cats, vals, colors):
        pct   = v / total
        sweep = pct * 360
        mid_r = np.deg2rad(start_deg - sweep / 2)
        cx    = 0.65 * R * np.cos(mid_r)
        cy    = 0.65 * R * np.sin(mid_r)
        ax.text(cx, cy,
                f"{lbl}\n{v/10000:.0f}万\n{pct*100:.1f}%",
                ha="center", va="center",
                fontsize=10.5, color="white", fontweight="bold", zorder=5)
        start_deg -= sweep

    ax.set_title("② カテゴリ別 構成", loc="left",
                 fontsize=12.5, fontweight="bold", color=C["text"], pad=4)


# ══════════════════════════════════════════════════════════════
# ② 曜日別平均売上（積み上げ棒）
# ══════════════════════════════════════════════════════════════
def plot_c2(ax, daily):
    order = ["火", "水", "木", "金", "土", "祝日"]
    DAY_C = {"火": C["c1"], "水": C["c1"], "木": C["c1"],
             "金": C["c4"], "土": C["c3"], "祝日": C["c5"]}
    wl, wd = defaultdict(list), defaultdict(list)
    for d in daily:
        if d["定休"] == "休":
            continue
        key = "祝日" if (d["曜日"] == "日" or d["祝日"]) else d["曜日"]
        if key in order:
            wl[key].append(d["昼食売上"]  / 10000)
            wd[key].append(d["夕食売上"] / 10000)
    al  = [np.mean(wl[w]) if wl[w] else 0 for w in order]
    ad  = [np.mean(wd[w]) if wd[w] else 0 for w in order]
    tot = [l + d for l, d in zip(al, ad)]
    top = max(tot) if tot else 1

    dark_ax(ax)
    for i, (w, l, d) in enumerate(zip(order, al, ad)):
        co = DAY_C[w]
        ax.bar(i, l, 0.72, color=co, alpha=0.95, edgecolor=C["bg"], lw=0.6, zorder=2)
        ax.bar(i, d, 0.72, bottom=l, color=co, alpha=0.40,
               edgecolor=C["bg"], lw=0.6, zorder=2)
        if l > top * 0.10:
            ax.text(i, l / 2, f"昼\n{l:.0f}万",
                    ha="center", va="center", fontsize=7.5,
                    color="white", fontweight="bold", zorder=3)
        if d > top * 0.10:
            ax.text(i, l + d / 2, f"夜\n{d:.0f}万",
                    ha="center", va="center", fontsize=7.5,
                    color=C["text"], fontweight="bold", zorder=3)
    for i, t in enumerate(tot):
        if t > 0:
            ax.text(i, t + top*0.04, f"{t:.0f}万",
                    ha="center", fontsize=8, color=C["text"], fontweight="bold")
    ax.set_xticks(range(len(order)))
    ax.set_xticklabels(order, fontsize=9.5, fontweight="bold")
    ax.set_ylim(0, top + 10)
    ax.set_ylabel("平均売上（万円）", color=C["sub"], fontsize=8.5, fontweight="bold")
    ax.set_title("③ 曜日別 平均売上", color=C["text"], fontsize=12.5,
                 fontweight="bold", pad=6)


# ══════════════════════════════════════════════════════════════
# ⑦ FOOD / DRINK ランキング（Top7）
# ══════════════════════════════════════════════════════════════
def plot_c7(ax_f, ax_d, shohin):
    def agg(kn, kq, ka):
        m = defaultdict(lambda: {"数量": 0, "金額": 0})
        for r in shohin:
            if r[kn]:
                m[r[kn]]["数量"] += r[kq]
                m[r[kn]]["金額"] += r[ka]
        top = sorted(m.items(), key=lambda x: x[1]["金額"], reverse=True)[:7]
        return ([x[0][:10] for x in top][::-1],
                [x[1]["金額"]/10000 for x in top][::-1],
                [x[1]["数量"] for x in top][::-1])

    def _draw(ax, names, amts, qtys, cm_name, title_color, title, unit, xlim_max):
        dark_ax(ax)
        ax.grid(axis="x", color=C["grid"], ls="--", lw=0.4, alpha=0.4)
        ax.grid(axis="y", visible=False)
        ax.spines["bottom"].set_visible(False)
        ax.spines["left"].set_visible(False)
        n    = len(names)
        cmap = matplotlib.colormaps.get_cmap(cm_name)
        grad = [cmap(0.35 + 0.65 * i / max(n-1, 1)) for i in range(n)]
        ax.barh(names, amts, color=grad, edgecolor=C["bg"], height=0.58)
        lbl_col   = "#0f172a"
        title_col = "black"
        for a, q, nm in zip(amts, qtys, names):
            ax.text(a + xlim_max * 0.02, names.index(nm),
                    f"{a:.1f}万  {q:,}{unit}",
                    va="center", fontsize=8, color=lbl_col, fontweight="bold")
        ax.set_xlim(0, xlim_max)
        ax.set_yticks(range(len(names)))
        ax.set_yticklabels(names, fontsize=8.5, color=lbl_col, fontweight="bold")
        ax.tick_params(axis="y", length=0, colors=lbl_col)
        ax.set_title(title, color=title_col, fontsize=12.5,
                     fontweight="bold", pad=5)

    fn, fa, fq = agg("F商品名", "F数量", "F金額")
    dn, da, dq = agg("D商品名", "D数量", "D金額")
    f_xlim = (max(fa) if fa else 0) + 50
    d_xlim = (max(da) if da else 0) + 10
    _draw(ax_f, fn, fa, fq, "YlOrRd", C["c4"], "④ FOOD ランキング Top7",  "食", f_xlim)
    _draw(ax_d, dn, da, dq, "Blues",  C["c1"], "⑤ DRINK ランキング Top7", "杯", d_xlim)


# ══════════════════════════════════════════════════════════════
# ③ 支払方法別（100%積み上げ横棒・下右エリア）
# ══════════════════════════════════════════════════════════════
def plot_c3(ax, daily):
    _lbls   = ["現金", "JCB", "千葉銀行", "アクアコイン", "PayPay", "売掛金"]
    _colors = [C["c1"], C["c3"], C["c4"], C["c2"], C["c5"], C["c7"]]
    _vals   = [sum(d[k] for d in daily) for k in ["現金", "JCB", "千葉銀行", "アクアコイン", "PayPay"]]
    _vals  += [sum(d["ふるさと納税"] + d["売掛金"] for d in daily)]
    order   = sorted(range(len(_vals)), key=lambda i: _vals[i], reverse=True)
    lbls    = [_lbls[i] for i in order]
    vals    = [_vals[i] for i in order]
    colors  = [_colors[i] for i in order]
    total   = sum(vals)
    pcts    = [v / total * 100 for v in vals]

    ax.set_facecolor(C["card"])
    for sp in ("top", "right", "left"):
        ax.spines[sp].set_visible(False)
    ax.spines["bottom"].set_edgecolor(C["grid"])
    ax.set_yticks([])
    ax.tick_params(axis="x", colors=C["text"], labelsize=9, length=3, pad=3)
    ax.grid(axis="x", color=C["grid"], ls="--", lw=0.4, alpha=0.35, zorder=0)

    THRESH_IN = 10.0   # %以上はバー内側、未満は引き出し線
    BAR_H     = 0.82
    left      = 0.0
    for lbl, v, pct, co in zip(lbls, vals, pcts, colors):
        ax.barh(0, pct, left=left, color=co, edgecolor=C["bg"],
                linewidth=1.2, height=BAR_H, zorder=2)
        mid = left + pct / 2
        if pct >= THRESH_IN:
            ax.text(mid, 0, f"{lbl}\n{v/10000:.0f}万\n{pct:.1f}%",
                    ha="center", va="center", fontsize=9.5,
                    color="white", fontweight="bold", zorder=5)
        else:
            top = BAR_H / 2
            ax.plot([mid, mid], [top, top + 0.07],
                    color=co, lw=0.9, alpha=0.7, zorder=4)
            ax.text(mid, top + 0.09, f"{lbl}\n{v/10000:.0f}万\n{pct:.1f}%",
                    ha="center", va="bottom", fontsize=8.5,
                    color=co, fontweight="bold", zorder=5)
        left += pct

    ax.set_xlim(0, 100)
    ax.set_ylim(-0.55, 1.10)
    ax.xaxis.set_major_formatter(plt.FuncFormatter(lambda v, _: f"{v:.0f}%"))
    sub_title(ax, "⑦ 支払方法別 構成（100%積み上げ）")


# ══════════════════════════════════════════════════════════════
# ⑧ 客単価 日次推移
# ══════════════════════════════════════════════════════════════
def plot_c8(ax, daily):
    op = [d for d in daily
          if d["定休"] != "休" and d["昼食客数"] > 0 and d["夕食客数"] > 0]
    days   = [d["日"] for d in op]
    l_unit = [d["昼食売上"] / d["昼食客数"] for d in op]
    d_unit = [d["夕食売上"] / d["夕食客数"] for d in op]
    l_avg, d_avg = np.mean(l_unit), np.mean(d_unit)

    dark_ax(ax)
    ax.fill_between(days, l_unit, d_unit,
                    where=[l > d for l, d in zip(l_unit, d_unit)],
                    alpha=0.18, color=C["c4"])
    ax.fill_between(days, l_unit, d_unit,
                    where=[d >= l for l, d in zip(l_unit, d_unit)],
                    alpha=0.18, color=C["c1"])
    neon_line(ax, days, l_unit, C["c4"], lw=1.8, ms=3.5, label="昼食 客単価")
    neon_line(ax, days, d_unit, C["c1"], lw=1.8, ms=3.5, marker="s", label="夕食 客単価")
    for co, avg, lbl in [(C["c4"], l_avg, f"昼平均 {l_avg:,.0f}"),
                          (C["c1"], d_avg, f"夜平均 {d_avg:,.0f}")]:
        for w, a in [(3, 0.06), (1.5, 0.12)]:
            ax.axhline(avg, color=co, lw=w, alpha=a, ls=":")
        ax.axhline(avg, color=co, lw=0.9, ls=":", alpha=0.6, label=lbl)
    ax.yaxis.set_major_formatter(plt.FuncFormatter(lambda v, _: f"{v:,.0f}"))
    ax.set_xlim(min(days) - 0.5, max(days) + 0.5)
    ax.set_ylabel("客単価（円）", color=C["sub"], fontsize=9, fontweight="bold")
    ax.legend(facecolor=C["card"], edgecolor=C["grid"],
              labelcolor=C["text"], fontsize=9, framealpha=0.85,
              loc="upper left", borderpad=0.5, ncol=2)
    set_day_xticks(ax, op)
    sub_title(ax, "⑥ 客単価 日次推移")


# ══════════════════════════════════════════════════════════════
# メイン描画
# ══════════════════════════════════════════════════════════════
def build_dashboard(year, month, daily, shohin):
    fig = plt.figure(figsize=(22, 13.5), facecolor=C["bg"])

    fig.text(0.5, 0.977,
             f"{year}年{month}月  営業日報 ダッシュボード",
             ha="center", va="top", fontsize=24, fontweight="bold",
             color=C["c1"])

    # 上段: KPI（row0）+ ①②（row1）を密着配置
    gs_top = gridspec.GridSpec(
        2, 12, figure=fig,
        height_ratios=[0.55, 3.4],
        hspace=0.06, wspace=0.24,
        top=0.920, bottom=0.570,
        left=0.042, right=0.972,
    )
    # 下段: ③④⑤（row0）+ ⑥⑦（row1）を密着配置
    gs_bot = gridspec.GridSpec(
        2, 12, figure=fig,
        height_ratios=[3.8, 3.1],
        hspace=0.22, wspace=0.24,
        top=0.508, bottom=0.048,
        left=0.042, right=0.972,
    )

    # ── KPI カード ─────────────────────────────────────────────
    op = [d for d in daily if d["定休"] != "休"]
    total_sales = sum(d["総売上高"] for d in daily)
    total_pax   = sum(d["昼食客数"] + d["夕食客数"] for d in op)
    total_lunch = sum(d["昼食売上"] for d in op)
    total_din   = sum(d["夕食売上"] for d in op)
    avg_unit    = (total_lunch + total_din) / total_pax if total_pax else 0
    kpis = [
        (gs_top[0, 0:3],   "総売上高（税込）",
         f"{total_sales/10000:.0f}万円",                          C["c1"]),
        (gs_top[0, 3:5],   "総来客数",
         f"{total_pax:,}名",                                      C["c3"]),
        (gs_top[0, 5:7],   "平均客単価",
         f"{avg_unit:,.0f}",                                      C["c4"]),
        (gs_top[0, 7:10],  "平均売上高",
         f"{total_sales/len(op)/10000:.0f}万円" if op else "―",   C["c2"]),
        (gs_top[0, 10:12], "稼働日数",
         f"{len(op)}日",                                          C["c7"]),
    ]
    for spec, lbl, val, co in kpis:
        draw_kpi(fig.add_subplot(spec), lbl, val, co)

    # ── ① 日次売上トレンド（左8/12）────────────────────────
    ax1 = fig.add_subplot(gs_top[1, 0:8])
    plot_c1(ax1, daily)

    # ── ② カテゴリ別構成（右4/12）───────────────────────────
    ax4 = fig.add_subplot(gs_top[1, 8:12])
    plot_c4(ax4, daily)
    # ② のみKPIとの間を空ける（上端を下げる）
    p = ax4.get_position()
    ax4.set_position([p.x0, p.y0, p.width, p.height - 0.04])

    # ── ③ 曜日別（cols0-2）──────────────────────────────────
    ax2 = fig.add_subplot(gs_bot[0, 0:3])
    plot_c2(ax2, daily)

    # ── ④⑤ FOOD[3:7] / DRINK[8:12] ─────────────────────────
    ax7f = fig.add_subplot(gs_bot[0, 3:7])
    ax7d = fig.add_subplot(gs_bot[0, 8:12])
    plot_c7(ax7f, ax7d, shohin)
    # ④ 左端を③ラベルと重ならないようオフセット
    p = ax7f.get_position()
    ax7f.set_position([p.x0 + 0.030, p.y0, p.width, p.height])

    # ── ⑥ 客単価（左8/12）──────────────────────────────────
    ax8 = fig.add_subplot(gs_bot[1, 0:8])
    plot_c8(ax8, daily)

    # ── ⑦ 支払方法別（右4/12）──────────────────────────────
    ax3 = fig.add_subplot(gs_bot[1, 8:12])
    plot_c3(ax3, daily)

    return fig


# ══════════════════════════════════════════════════════════════
# メイン
# ══════════════════════════════════════════════════════════════
def main():
    parser = argparse.ArgumentParser(description="日次データ → 1枚ダッシュボード PNG")
    parser.add_argument("year",  type=int)
    parser.add_argument("month", type=int)
    args = parser.parse_args()

    out = OUT_DIR / f"dashboard_{args.year}_{args.month}.png"
    print(f"データ読み込み中: daily_{args.year}_{args.month}.xlsx")
    daily, shohin = load_data(args.year, args.month)
    print(f"  日次: {len(daily)}日  商品別: {len(shohin)}行")

    print("ダッシュボード生成中...")
    fig = build_dashboard(args.year, args.month, daily, shohin)

    OUT_DIR.mkdir(parents=True, exist_ok=True)
    fig.savefig(out, dpi=150, bbox_inches="tight", facecolor=C["bg"])
    plt.close(fig)
    print(f"\n出力完了: {out}")


if __name__ == "__main__":
    main()
