#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
日次データExcel → 分析レポートExcel 生成スクリプト（ライトテーマ）

使い方:
    python create_report.py 2026 5
    → frontend/public/data/report_2026_5.xlsx を出力

チャート構成（ダッシュボードと同一）:
  ① 日次売上トレンド        折れ線 + 面
  ② カテゴリ別 構成         円グラフ
  ③ 曜日別 平均売上         積み上げ棒（昼・夜）
  ④⑤ FOOD・DRINK ランキング 横棒（金額+数量）
  ⑥ 客単価 日次推移         折れ線
  ⑦ 支払方法別 構成         100%積み上げ横棒
"""
import argparse
import io
from pathlib import Path
from collections import defaultdict
import numpy as np
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage
import matplotlib.font_manager as fm

ROOT    = Path(__file__).parent.parent
IN_DIR  = ROOT / "frontend" / "public" / "data"
OUT_DIR = ROOT / "frontend" / "public" / "data"

_avail = {f.name for f in fm.fontManager.ttflist}
JP = next((f for f in ["Yu Gothic", "Meiryo", "MS Gothic"] if f in _avail), "sans-serif")
plt.rcParams.update({
    "font.family":        JP,
    "axes.unicode_minus": False,
    "axes.labelpad":      6,
    "xtick.major.pad":    5,
    "ytick.major.pad":    5,
})

# ライトテーマ カラーパレット
C = {
    "bg":   "#f0f4f8", "card": "#ffffff",
    "c1":   "#0284c7", "c2":   "#7c3aed",
    "c3":   "#059669", "c4":   "#d97706",
    "c5":   "#e11d48", "c6":   "#ca8a04",
    "c7":   "#0891b2", "c8":   "#be123c",
    "grid": "#cbd5e1", "text": "#0f172a", "sub":  "#475569",
}

# Excel スタイル（ライトテーマ）
HDR_FILL = PatternFill("solid", fgColor="0284C7")
HDR_FONT = Font(bold=True, color="FFFFFF", size=11)
TTL_FONT = Font(bold=True, color="0284C7", size=14)
VAL_FONT = Font(bold=True, color="0F172A", size=12)
THIN     = Side(style="thin", color="CBD5E1")
BORDER   = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER   = Alignment(horizontal="center", vertical="center")
LEFT     = Alignment(horizontal="left",   vertical="center")


# ══════════════════════════════════════════════════════════════
# 描画ヘルパー
# ══════════════════════════════════════════════════════════════
def theme_ax(ax, fig=None):
    if fig:
        fig.patch.set_facecolor(C["bg"])
    ax.set_facecolor(C["card"])
    ax.tick_params(colors=C["text"], labelsize=9, length=3,
                   direction="out", pad=4, width=0.8)
    ax.spines["top"].set_visible(False)
    ax.spines["right"].set_visible(False)
    ax.spines["left"].set_edgecolor(C["grid"])
    ax.spines["bottom"].set_edgecolor(C["grid"])
    ax.grid(color=C["grid"], ls="--", lw=0.6, alpha=0.7, axis="y", zorder=0)


def clean_line(ax, xs, ys, color, lw=2.2, ms=5, marker="o", label="", zorder=3):
    ax.plot(xs, ys, color=color, lw=lw, marker=marker, ms=ms,
            markerfacecolor=C["bg"], markeredgecolor=color,
            markeredgewidth=1.8, label=label,
            solid_capstyle="round", zorder=zorder)


def set_day_xticks(ax, op_days):
    xs     = [d["日"] for d in op_days]
    labels = [f"{d['日']}\n{d['曜日']}" for d in op_days]
    ax.set_xticks(xs)
    ax.set_xticklabels(labels, fontsize=7)
    ax.tick_params(axis="x", length=3, pad=1)


def to_img(fig, dpi=150):
    buf = io.BytesIO()
    fig.savefig(buf, format="png", dpi=dpi, bbox_inches="tight",
                facecolor=fig.get_facecolor())
    buf.seek(0)
    plt.close(fig)
    return buf


# ══════════════════════════════════════════════════════════════
# データ読み込み
# ══════════════════════════════════════════════════════════════
def load_data(year: int, month: int):
    path = IN_DIR / f"daily_{year}_{month}.xlsx"
    if not path.exists():
        raise FileNotFoundError(f"データファイルが見つかりません: {path}")
    wb  = openpyxl.load_workbook(path)
    ws1 = wb.active
    keys = [
        "日付", "日", "曜日", "定休", "祝日",
        "昼天気", "夜天気",
        "総売上高", "純売上高",
        "現金", "JCB", "千葉銀行", "アクアコイン", "PayPay", "ふるさと納税", "売掛金",
        "FOOD", "DRINK", "売店", "その他",
        "昼食客数", "夕食客数", "昼食売上", "夕食売上",
    ]
    daily = []
    for r in range(3, ws1.max_row + 1):
        v = [ws1.cell(row=r, column=c).value for c in range(1, len(keys) + 1)]
        if v[0] is None or str(v[0]).startswith("合"):
            continue
        row = dict(zip(keys, v))
        for k in keys[7:]:
            row[k] = int(row[k] or 0)
        daily.append(row)

    ws2    = wb["商品別"]
    s_keys = ["日付", "日", "曜日", "順位",
              "F商品名", "F単価", "F数量", "F金額",
              "D商品名", "D単価", "D数量", "D金額"]
    shohin = []
    for r in range(3, ws2.max_row + 1):
        v = [ws2.cell(row=r, column=c).value for c in range(1, 13)]
        if v[0] is None or str(v[0]).startswith("合"):
            continue
        row = dict(zip(s_keys, v))
        for k in ["F単価", "F数量", "F金額", "D単価", "D数量", "D金額"]:
            row[k] = int(row[k] or 0)
        shohin.append(row)
    return daily, shohin


# ══════════════════════════════════════════════════════════════
# チャート定義
# ══════════════════════════════════════════════════════════════

def chart_1(daily):
    """① 日次売上トレンド"""
    op     = [d for d in daily if d["定休"] != "休"]
    xs     = [d["日"] for d in op]
    total  = [d["総売上高"] / 10000 for d in op]
    lunch  = [d["昼食売上"]  / 10000 for d in op]
    dinner = [d["夕食売上"]  / 10000 for d in op]
    avg    = np.mean(total)
    max_i  = int(np.argmax(total))

    fig, ax = plt.subplots(figsize=(12, 5))
    theme_ax(ax, fig)

    ax.fill_between(xs, total,  alpha=0.28, color=C["c1"], zorder=1)
    ax.fill_between(xs, lunch,  alpha=0.22, color=C["c4"], zorder=1)
    ax.fill_between(xs, dinner, alpha=0.22, color=C["c2"], zorder=1)

    clean_line(ax, xs, total,  C["c1"], lw=2.4, ms=4.5, label="総売上高")
    clean_line(ax, xs, lunch,  C["c4"], lw=1.7, ms=3.5, marker="^", label="昼食売上")
    clean_line(ax, xs, dinner, C["c2"], lw=1.7, ms=3.5, marker="s", label="夕食売上")

    ax.axhline(avg, color=C["c3"], lw=1.5, ls="--", alpha=0.8,
               label=f"平均 {avg:.0f}万")

    ax.annotate(f"MAX {xs[max_i]}日\n{total[max_i]:.0f}万",
                xy=(xs[max_i], total[max_i]),
                xytext=(xs[max_i], total[max_i] + max(total) * 0.06),
                ha="center", fontsize=8.5, color=C["c4"], fontweight="bold",
                bbox=dict(boxstyle="round,pad=0.3", fc="white",
                          ec=C["c4"], alpha=0.9, lw=1))

    ax.set_ylabel("売上（万円）", color=C["sub"], fontsize=9)
    ax.set_ylim(0, max(total) * 1.25)
    ax.set_xlim(min(xs) - 0.5, max(xs) + 0.5)
    ax.legend(facecolor="white", edgecolor=C["grid"],
              labelcolor=C["text"], fontsize=9.5, framealpha=0.9,
              loc="upper left", borderpad=0.7)
    set_day_xticks(ax, op)
    ax.set_title("① 日次売上トレンド",
                 color=C["text"], fontsize=14, fontweight="bold", pad=14)
    fig.tight_layout()
    return to_img(fig)


def chart_2(daily):
    """② カテゴリ別 構成（円グラフ・FOOD/DRINK/その他）"""
    cats   = ["FOOD", "DRINK", "その他"]
    colors = [C["c4"], C["c1"], C["c2"]]
    raw    = {k: sum(d[k] for d in daily) for k in ["FOOD", "DRINK", "売店", "その他"]}
    vals   = [raw["FOOD"], raw["DRINK"], raw["売店"] + raw["その他"]]
    total  = sum(vals)

    fig, ax = plt.subplots(figsize=(8, 6))
    fig.patch.set_facecolor(C["bg"])
    ax.set_facecolor(C["bg"])
    ax.set_aspect("equal")

    ax.pie(vals, colors=colors, startangle=90,
           wedgeprops=dict(edgecolor=C["bg"], linewidth=2.5),
           counterclock=False)

    start_deg = 90.0
    for lbl, v, co in zip(cats, vals, colors):
        pct   = v / total
        sweep = pct * 360
        mid_r = np.deg2rad(start_deg - sweep / 2)
        cx    = 0.60 * np.cos(mid_r)
        cy    = 0.60 * np.sin(mid_r)
        ax.text(cx, cy,
                f"{lbl}\n{v/10000:.0f}万\n{pct*100:.1f}%",
                ha="center", va="center",
                fontsize=10, color="#03071e", fontweight="bold", zorder=5)
        start_deg -= sweep

    ax.set_title("② カテゴリ別 構成",
                 color=C["text"], fontsize=14, fontweight="bold")
    fig.tight_layout()
    return to_img(fig)


def chart_3(daily):
    """③ 曜日別 平均売上（昼夜積み上げ棒）"""
    order = ["火", "水", "木", "金", "土", "祝日"]
    DAY_C = {"火": C["c1"], "水": C["c1"], "木": C["c1"],
             "金": C["c4"], "土": C["c3"], "祝日": C["c5"]}
    wl, wd = defaultdict(list), defaultdict(list)
    for d in daily:
        if d["定休"] == "休":
            continue
        key = "祝日" if (d["曜日"] == "日" or d["祝日"]) else d["曜日"]
        if key in order:
            wl[key].append(d["昼食売上"]  / 10000)
            wd[key].append(d["夕食売上"] / 10000)
    al  = [np.mean(wl[w]) if wl[w] else 0 for w in order]
    ad  = [np.mean(wd[w]) if wd[w] else 0 for w in order]
    tot = [l + d for l, d in zip(al, ad)]
    top = max(tot) if tot else 1

    fig, ax = plt.subplots(figsize=(9, 5.5))
    theme_ax(ax, fig)
    for i, (w, l, d) in enumerate(zip(order, al, ad)):
        co = DAY_C[w]
        ax.bar(i, l, 0.52, color=co, alpha=0.95, edgecolor=C["bg"], lw=0.8, zorder=2)
        ax.bar(i, d, 0.52, bottom=l, color=co, alpha=0.40,
               edgecolor=C["bg"], lw=0.8, zorder=2)
    for i, (t, al_, ad_) in enumerate(zip(tot, al, ad)):
        if t > 0:
            ax.text(i, t + top * 0.03, f"{t:.0f}万",
                    ha="center", fontsize=8, color=C["text"], fontweight="bold")
    ax.set_xticks(range(len(order)))
    ax.set_xticklabels(order, fontsize=11)
    ax.set_ylim(0, top * 1.32)
    ax.set_ylabel("平均売上（万円）", color=C["sub"], fontsize=9)
    p1 = mpatches.Patch(color=C["c1"], alpha=0.95, label="昼食売上（濃色）")
    p2 = mpatches.Patch(color=C["c1"], alpha=0.40, label="夕食売上（淡色）")
    ax.legend(handles=[p1, p2], facecolor="white", edgecolor=C["grid"],
              labelcolor=C["text"], fontsize=9, framealpha=0.9, loc="upper left")
    ax.set_title("③ 曜日別 平均売上",
                 color=C["text"], fontsize=14, fontweight="bold", pad=14)
    fig.tight_layout()
    return to_img(fig)


def chart_45(shohin):
    """④⑤ FOOD・DRINK ランキング Top7（横棒グラフ）"""
    def agg(kn, kq, ka):
        m = defaultdict(lambda: {"数量": 0, "金額": 0})
        for r in shohin:
            if r[kn]:
                m[r[kn]]["数量"] += r[kq]
                m[r[kn]]["金額"] += r[ka]
        top = sorted(m.items(), key=lambda x: x[1]["金額"], reverse=True)[:7]
        return ([x[0][:10] for x in top][::-1],
                [x[1]["金額"]/10000 for x in top][::-1],
                [x[1]["数量"] for x in top][::-1])

    fn, fa, fq = agg("F商品名", "F数量", "F金額")
    dn, da, dq = agg("D商品名", "D数量", "D金額")

    fig, (ax_f, ax_d) = plt.subplots(1, 2, figsize=(13, 5.5))
    fig.patch.set_facecolor(C["bg"])

    def _draw(ax, names, amts, qtys, cm_name, title, unit, xlim_max):
        theme_ax(ax)
        ax.grid(axis="x", color=C["grid"], ls="--", lw=0.4, alpha=0.5)
        ax.grid(axis="y", visible=False)
        ax.spines["bottom"].set_visible(False)
        ax.spines["left"].set_visible(False)
        n    = len(names)
        cmap = matplotlib.colormaps.get_cmap(cm_name)
        grad = [cmap(0.35 + 0.65 * i / max(n-1, 1)) for i in range(n)]
        ax.barh(names, amts, color=grad, edgecolor=C["bg"], height=0.52)
        for a, q, nm in zip(amts, qtys, names):
            ax.text(a + xlim_max * 0.02, names.index(nm),
                    f"{a:.1f}万  {q:,}{unit}",
                    va="center", fontsize=7.5, color=C["text"], fontweight="bold")
        ax.set_xlim(0, xlim_max)
        ax.set_yticks(range(len(names)))
        ax.set_yticklabels(names, fontsize=8, color=C["text"], fontweight="bold")
        ax.tick_params(axis="y", length=0)
        ax.set_title(title, color=C["text"], fontsize=12, fontweight="bold", pad=8)

    _draw(ax_f, fn, fa, fq, "YlOrRd", "④ FOOD ランキング Top7",  "個", 350)
    _draw(ax_d, dn, da, dq, "Blues",  "⑤ DRINK ランキング Top7", "杯",  60)
    fig.tight_layout()
    return to_img(fig)


def chart_6(daily):
    """⑥ 客単価 日次推移（折れ線）"""
    op = [d for d in daily
          if d["定休"] != "休" and d["昼食客数"] > 0 and d["夕食客数"] > 0]
    days   = [d["日"] for d in op]
    l_unit = [d["昼食売上"] / d["昼食客数"] for d in op]
    d_unit = [d["夕食売上"] / d["夕食客数"] for d in op]
    l_avg, d_avg = np.mean(l_unit), np.mean(d_unit)

    fig, ax = plt.subplots(figsize=(12, 5))
    theme_ax(ax, fig)
    ax.fill_between(days, l_unit, d_unit,
                    where=[l > d for l, d in zip(l_unit, d_unit)],
                    alpha=0.18, color=C["c4"])
    ax.fill_between(days, l_unit, d_unit,
                    where=[d >= l for l, d in zip(l_unit, d_unit)],
                    alpha=0.18, color=C["c1"])
    clean_line(ax, days, l_unit, C["c4"], lw=2.2, ms=5, label="昼食 客単価")
    clean_line(ax, days, d_unit, C["c1"], lw=2.2, ms=5, marker="s", label="夕食 客単価")
    for co, avg, lbl in [(C["c4"], l_avg, f"昼平均 {l_avg:,.0f}"),
                          (C["c1"], d_avg, f"夜平均 {d_avg:,.0f}")]:
        ax.axhline(avg, color=co, lw=1.5, ls=":", alpha=0.8, label=lbl)
    ax.yaxis.set_major_formatter(plt.FuncFormatter(lambda v, _: f"{v:,.0f}"))
    ax.set_ylabel("客単価（円）", color=C["sub"], fontsize=9)
    ax.set_xlim(min(days) - 0.5, max(days) + 0.5)
    ax.legend(facecolor="white", edgecolor=C["grid"],
              labelcolor=C["text"], fontsize=9.5, framealpha=0.9,
              loc="upper left", borderpad=0.7, ncol=2)
    set_day_xticks(ax, op)
    ax.set_title("⑥ 客単価 日次推移",
                 color=C["text"], fontsize=14, fontweight="bold", pad=14)
    fig.tight_layout()
    return to_img(fig)


def chart_7(daily):
    """⑦ 支払方法別 構成（100%積み上げ横棒）"""
    _lbls   = ["現金", "JCB", "千葉銀行", "アクアコイン", "PayPay", "売掛金"]
    _colors = [C["c1"], C["c3"], C["c4"], C["c2"], C["c5"], C["c7"]]
    _vals   = [sum(d[k] for d in daily) for k in ["現金", "JCB", "千葉銀行", "アクアコイン", "PayPay"]]
    _vals  += [sum(d["ふるさと納税"] + d["売掛金"] for d in daily)]
    order   = sorted(range(len(_vals)), key=lambda i: _vals[i], reverse=True)
    lbls    = [_lbls[i] for i in order]
    vals    = [_vals[i] for i in order]
    colors  = [_colors[i] for i in order]
    total   = sum(vals)
    pcts    = [v / total * 100 for v in vals]

    fig, ax = plt.subplots(figsize=(10, 3.5))
    fig.patch.set_facecolor(C["bg"])
    ax.set_facecolor(C["card"])
    for sp in ("top", "right", "left"):
        ax.spines[sp].set_visible(False)
    ax.spines["bottom"].set_edgecolor(C["grid"])
    ax.set_yticks([])
    ax.tick_params(axis="x", colors=C["text"], labelsize=9, length=3, pad=3)
    ax.grid(axis="x", color=C["grid"], ls="--", lw=0.4, alpha=0.5, zorder=0)

    THRESH_IN = 10.0
    BAR_H     = 0.75
    left      = 0.0
    for lbl, v, pct, co in zip(lbls, vals, pcts, colors):
        ax.barh(0, pct, left=left, color=co, edgecolor=C["bg"],
                linewidth=1.2, height=BAR_H, zorder=2)
        mid = left + pct / 2
        if pct >= THRESH_IN:
            ax.text(mid, 0, f"{lbl}\n{v/10000:.0f}万\n{pct:.1f}%",
                    ha="center", va="center", fontsize=8.5,
                    color="#03071e", fontweight="bold", zorder=5)
        else:
            top_y = BAR_H / 2
            ax.plot([mid, mid], [top_y, top_y + 0.08],
                    color=co, lw=0.9, alpha=0.7, zorder=4)
            ax.text(mid, top_y + 0.10, f"{lbl}\n{v/10000:.0f}万\n{pct:.1f}%",
                    ha="center", va="bottom", fontsize=7.5,
                    color=co, fontweight="bold", zorder=5)
        left += pct

    ax.set_xlim(0, 100)
    ax.set_ylim(-0.55, 1.15)
    ax.xaxis.set_major_formatter(plt.FuncFormatter(lambda v, _: f"{v:.0f}%"))
    ax.set_title("⑦ 支払方法別 構成（100%積み上げ）",
                 color=C["text"], fontsize=14, fontweight="bold", pad=10)
    fig.tight_layout()
    return to_img(fig)


# ══════════════════════════════════════════════════════════════
# レポートExcel 書き出し
# ══════════════════════════════════════════════════════════════
CHART_META = [
    ("① 日次売上トレンド",         (12, 5.0), "折れ線 + 面"),
    ("② カテゴリ別 構成",          ( 8, 6.0), "円グラフ"),
    ("③ 曜日別 平均売上",          ( 9, 5.5), "積み上げ棒グラフ"),
    ("④⑤ FOOD・DRINK ランキング",  (13, 5.5), "横棒グラフ"),
    ("⑥ 客単価 日次推移",          (12, 5.0), "折れ線グラフ"),
    ("⑦ 支払方法別 構成",          (10, 3.5), "100%積み上げ横棒"),
]


def write_index(ws, year, month, daily):
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 2

    ws.merge_cells("B1:L1")
    t = ws["B1"]
    t.value = f"{year}年{month}月 営業日報 分析レポート"
    t.font = TTL_FONT
    t.fill = PatternFill("solid", fgColor="EFF6FF")
    t.alignment = CENTER
    t.border = Border(left=Side(style="medium", color="0284C7"),
                      right=Side(style="medium", color="0284C7"),
                      top=Side(style="medium", color="0284C7"),
                      bottom=Side(style="medium", color="0284C7"))
    ws.row_dimensions[1].height = 36

    op = [d for d in daily if d["定休"] != "休"]
    total_pax = sum(d["昼食客数"] + d["夕食客数"] for d in op)
    avg_unit  = (sum(d["昼食売上"] + d["夕食売上"] for d in op) / total_pax
                 if total_pax else 0)
    kpis = [
        ("総売上高（税込）", f"{sum(d['総売上高'] for d in daily)/10000:.0f}万円", C["c1"]),
        ("総来客数",         f"{total_pax:,}名",                                   C["c3"]),
        ("平均客単価",       f"{avg_unit:,.0f}円",                                 C["c4"]),
        ("FOOD売上",         f"{sum(d['FOOD'] for d in daily)/10000:.0f}万円",     C["c2"]),
        ("稼働日数",         f"{len(op)}日",                                        C["c7"]),
    ]
    for i, (lbl, val, color) in enumerate(kpis):
        col = 2 + i * 2
        lc = get_column_letter(col)
        vc = get_column_letter(col + 1)
        ws.column_dimensions[lc].width = 14
        ws.column_dimensions[vc].width = 14
        hex_c = color.lstrip("#")

        cl = ws.cell(row=3, column=col, value=lbl)
        cl.font = Font(color=hex_c, size=10, bold=True)
        cl.fill = PatternFill("solid", fgColor="F8FAFC")
        cl.alignment = CENTER
        cl.border = Border(bottom=Side(style="medium", color=hex_c))
        ws.merge_cells(f"{lc}3:{vc}3")

        cv = ws.cell(row=4, column=col, value=val)
        cv.font = Font(color=hex_c, size=14, bold=True)
        cv.fill = PatternFill("solid", fgColor="FFFFFF")
        cv.alignment = CENTER
        cv.border = Border(bottom=Side(style="thin", color="CBD5E1"))
        ws.merge_cells(f"{lc}4:{vc}4")
    ws.row_dimensions[3].height = 22
    ws.row_dimensions[4].height = 32

    ws.merge_cells("B6:L6")
    s = ws["B6"]
    s.value = "■ 分析チャート一覧"
    s.font = Font(bold=True, color="0284C7", size=11)
    s.fill = PatternFill("solid", fgColor="EFF6FF")
    s.alignment = LEFT
    ws.row_dimensions[6].height = 22

    for ci, lbl in enumerate(["#", "チャート名", "グラフ種類"], 1):
        c = ws.cell(row=7, column=ci + 1, value=lbl)
        c.font = HDR_FONT; c.fill = HDR_FILL
        c.alignment = CENTER; c.border = BORDER
    ws.column_dimensions["B"].width = 5
    ws.column_dimensions["C"].width = 30
    ws.column_dimensions["D"].width = 26

    for i, (name, _, chart_type) in enumerate(CHART_META, 1):
        r = 7 + i
        for ci, (val, fnt) in enumerate([
            (i,          VAL_FONT),
            (name,       VAL_FONT),
            (chart_type, Font(color="475569", size=10)),
        ], 2):
            c = ws.cell(row=r, column=ci, value=val)
            c.font = fnt
            c.fill = PatternFill("solid", fgColor="F8FAFC")
            c.alignment = LEFT if ci == 3 else CENTER
            c.border = BORDER
        ws.row_dimensions[r].height = 20


def write_chart_sheet(wb, sheet_name: str, img_buf, w_in: float, h_in: float):
    ws = wb.create_sheet(sheet_name)
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = "0284C7"
    img = XLImage(img_buf)
    img.width  = int(w_in * 96)
    img.height = int(h_in * 96)
    ws.add_image(img, "A1")


def write_report(year: int, month: int, imgs, daily):
    wb = openpyxl.Workbook()
    ws_idx = wb.active
    ws_idx.title = "サマリー"
    ws_idx.sheet_properties.tabColor = "0284C7"
    write_index(ws_idx, year, month, daily)

    for img, (name, (w, h), _) in zip(imgs, CHART_META):
        write_chart_sheet(wb, name, img, w, h)

    out = OUT_DIR / f"report_{year}_{month}.xlsx"
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    return out


# ══════════════════════════════════════════════════════════════
# メイン
# ══════════════════════════════════════════════════════════════
def main():
    parser = argparse.ArgumentParser(description="日次データ → 分析レポートExcel（ライトテーマ）")
    parser.add_argument("year",  type=int)
    parser.add_argument("month", type=int)
    args = parser.parse_args()

    print(f"データ読み込み中: daily_{args.year}_{args.month}.xlsx")
    daily, shohin = load_data(args.year, args.month)
    print(f"  日次: {len(daily)}日  商品別: {len(shohin)}行")

    print("チャート生成中...")
    chart_fns = [
        (chart_1,  daily),
        (chart_2,  daily),
        (chart_3,  daily),
        (chart_45, shohin),
        (chart_6,  daily),
        (chart_7,  daily),
    ]
    imgs = []
    for i, (fn, arg) in enumerate(chart_fns, 1):
        imgs.append(fn(arg))
        print(f"  {i}/{len(chart_fns)} 完了")

    out = write_report(args.year, args.month, imgs, daily)
    print(f"\n出力完了: {out}")
    print(f"  シート構成: サマリー + {len(CHART_META)}チャートシート")


if __name__ == "__main__":
    main()
